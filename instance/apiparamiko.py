from flask import Flask, jsonify, request
import paramiko
import openpyxl

app = Flask(__name__)

def get_linux_credentials(data):
    return (
        data.get('host'),
        data.get('port'),
        data.get('username'),
        data.get('password')
    )

def execute_linux_command(command, credentials):
    try:
        host, port, username, password = credentials
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, port=port, username=username, password=password)
        stdin, stdout, stderr = ssh.exec_command(command)
        result = stdout.read().decode('utf-8')
        exit_status = stdout.channel.recv_exit_status()
        ssh.close()
        return result, exit_status
    except paramiko.SSHException as ssh_exception:
        return f"SSH Exception: {str(ssh_exception)}", -1
    except Exception as e:
        return f"An error occurred: {str(e)}", -1

def export_health_to_excel(health_info):
    try:
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write server health information to Excel
        ws['A1'] = 'Metric'
        ws['B1'] = 'Value'
        ws['A2'] = 'CPU Usage'
        ws['B2'] = health_info.get('CPUUsage', 'N/A')
        ws['A3'] = 'Memory Usage'
        ws['B3'] = health_info.get('MemoryUsage', 'N/A')

        # Save Excel workbook to a Windows-compatible path with double backslashes
        excel_file_path = "C:\\Path\\New folder\\New XLS Worksheet.xls"
        wb.save(excel_file_path)

        return excel_file_path

    except Exception as e:
        raise Exception(f"Error exporting to Excel: {str(e)}")

@app.route('/automation/api/1.0/linux/server_health', methods=['POST'])
def linux_server_health():
    try:
        data = request.get_json()

        # Get Linux server credentials
        credentials = get_linux_credentials(data)

        # Execute Linux command to check server health
        command = "top -n 1 -b"
        result, exit_status = execute_linux_command(command, credentials)

        if exit_status == 0:
            # Parse the result for relevant metrics
            # For simplicity, assume you can extract CPU and Memory usage from the result
            cpu_usage = float(result.split('\n')[2].split()[1])
            memory_usage = float(result.split('\n')[3].split()[3].replace('%', ''))

            # Create a dictionary with the health information
            health_info = {'CPUUsage': cpu_usage, 'MemoryUsage': memory_usage}

            # Export health information to Excel
            excel_file_path = export_health_to_excel(health_info)

            return jsonify({'Status': 'Success', 'ExcelFilePath': excel_file_path})
        else:
            return jsonify({'error': result}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)

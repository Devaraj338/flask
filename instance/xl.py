import psutil
import paramiko
from flask import Flask, jsonify, request
import openpyxl

app = Flask(__name__)

def get_linux_credentials(data):
    return (
        data.get('host'),
        data.get('port'),
        data.get('username'),
        data.get('password')
    )

def execute_linux_command(command, credentials):
    try:
        host, port, username, password = credentials
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, port=port, username=username, password=password)
        stdin, stdout, stderr = ssh.exec_command(command)
        result = stdout.read().decode('utf-8')
        exit_status = stdout.channel.recv_exit_status()
        ssh.close()
        return result, exit_status
    except paramiko.SSHException as ssh_exception:
        return f"SSH Exception: {str(ssh_exception)}", -1
    except Exception as e:
        return f"An error occurred: {str(e)}", -1

@app.route('/linux/server_health', methods=['POST'])
def linux_server_health():
    try:
        data = request.get_json()
        
        # Get Linux server credentials
        credentials = get_linux_credentials(data)

        # Execute Linux command to check server health
        command = "top -n 1 -b"
        result, exit_status = execute_linux_command(command, credentials)

        if exit_status == 0:
            # Parse the result for relevant metrics
            # For simplicity, assume you can extract CPU and Memory usage from the result
            cpu_usage = float(result.split('\n')[2].split()[1])
            memory_usage = float(result.split('\n')[3].split()[3].replace('%', ''))

            # Create a dictionary with the Linux health information
            linux_health_info = {'CPUUsage': cpu_usage, 'MemoryUsage': memory_usage}

            # Get server health information
            server_credentials = (data.get('host'), data.get('username'), data.get('key_path'))
            server_health_info = get_server_health(server_credentials)

            # Combine Linux and server health information
            combined_health_info = {**linux_health_info, **server_health_info}

            # Export health information to Excel
            excel_file_path = export_health_to_excel(combined_health_info)

            return jsonify({'Status': 'Success', 'ExcelFilePath': excel_file_path})
        else:
            return jsonify({'error': result}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
def get_server_health(credentials):
    try:
        # Extract credentials
        host, username, key_path = credentials

        # Initialize a dictionary to store server health metrics
        health_info = {}

        # Get CPU information
        cpu_info = psutil.cpu_times_percent()
        health_info['CPUCount'] = psutil.cpu_count(logical=False)  # Number of physical CPUs
        health_info['CPUPercentage'] = psutil.cpu_percent()
        health_info['CPUInfo'] = cpu_info._asdict()

        # Get Memory information
        memory_info = psutil.virtual_memory()
        health_info['TotalRAM'] = round(memory_info.total / (1024 ** 3), 2)  # Convert to GB
        health_info['AvailableRAM'] = round(memory_info.available / (1024 ** 3), 2)  # Convert to GB
        health_info['MemoryPercentage'] = memory_info.percent

        # Get Swap information
        swap_info = psutil.swap_memory()
        health_info['SwapPercentage'] = swap_info.percent
        health_info['SwapUsedPercentage'] = (swap_info.used / swap_info.total) * 100

        # Get Disk information for each partition
        disk_partitions = psutil.disk_partitions()
        for partition in disk_partitions:
            partition_info = psutil.disk_usage(partition.mountpoint)
            partition_name = partition.device
            health_info[f'DiskUsagePercentage_{partition_name}'] = partition_info.percent

        # Get Hostname
        health_info['Hostname'] = host

        return health_info

    except Exception as e:
        raise Exception(f"Error getting server health information: {str(e)}")

# Function to export health information to Excel
def export_health_to_excel(health_info):
    try:
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write server health information to Excel
        ws['A1'] = 'Metric'
        ws['B1'] = 'Value'

        row = 2
        for metric, value in health_info.items():
            # Handle special cases, like CPU usage
            if isinstance(value, dict):
                for sub_metric, sub_value in value.items():
                    ws.cell(row=row, column=1, value=f'{metric} - {sub_metric}')
                    ws.cell(row=row, column=2, value=sub_value)
                    row += 1
            else:
                ws.cell(row=row, column=1, value=metric)
                ws.cell(row=row, column=2, value=value)
                row += 1

        # Save Excel workbook without specifying force_save
        excel_file_path = "C:\\server_health_check_report\\server_health_report.xlsx"  # Save in the same directory as the script

        wb.save(excel_file_path)

        return excel_file_path  # Return the file path

    except Exception as e:
        raise Exception(f"Error exporting to Excel: {str(e)}")

if __name__ == '__main__':
    app.run(debug=True)
